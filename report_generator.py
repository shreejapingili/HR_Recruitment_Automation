from openpyxl import Workbook
import os

# Create the reports folder automatically if it doesn't exist
os.makedirs("reports", exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "Summary"

# Add column headers
ws["A1"] = "Metric"
ws["B1"] = "Value"

# Append row data
ws.append(["Total Applications", 500])
ws.append(["Selected Candidates", 180])
ws.append(["Offers Released", 100])
ws.append(["Joined Candidates", 85])

# Save workbook
wb.save("reports/recruitment_report.xlsx")

print("Report Generated")
